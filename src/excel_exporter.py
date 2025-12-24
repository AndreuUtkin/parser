import pandas as pd
import logging
from typing import List
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from src.tariff import Tariff

logger = logging.getLogger(__name__)

#Экспорт тарифов excel
class ExcelExporter:
    
    def __init__(self, filename: str = "rialcom_tariffs.xlsx"):
        self.filename = filename
    
    def export_to_excel(self, tariffs: List[Tariff]):
        
        if not tariffs:
            logger.warning("Нет данных для экспорта")
            return
        
        logger.info(f"Экспорт {len(tariffs)} тарифов в Excel...")
        
        # Создаем DataFrame
        data = [tariff.to_dict() for tariff in tariffs]
        df = pd.DataFrame(data)
        
        
        # Создаем Excel writer
        with pd.ExcelWriter(self.filename, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Тарифы', index=False)
            
            # Получаем workbook и worksheet для форматирования
            workbook = writer.book
            worksheet = writer.sheets['Тарифы']
            
            # Применяем форматирование
            self._format_worksheet(worksheet, len(df))
            
            # Автонастройка ширины колонок
            self._auto_adjust_columns(worksheet)
            
            # Сохраняем
            workbook.save(self.filename)
        
        logger.info(f"Файл сохранен: {self.filename}")

    
    #Форматирование листа Excel
    def _format_worksheet(self, worksheet, data_rows: int):
        # Настройка стилей
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="444444", end_color="444444", fill_type="solid")
        center_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Форматирование заголовков
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = thin_border
        
        # Форматирование данных
        for row in worksheet.iter_rows(min_row=2, max_row=data_rows + 1, max_col=4):
            for cell in row:
                cell.border = thin_border
                
                # Выравнивание по центру для числовых колонок
                if cell.column in [2, 3, 4]: 
                    cell.alignment = center_alignment
                
                
    
    #Автонастройка ширины колонок
    def _auto_adjust_columns(self, worksheet):
        
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width